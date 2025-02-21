import base64
import json
import io
import fitz  # PyMuPDF
from typing import List, Tuple
from PIL import Image
import base64
import uuid
import os
from wasabi import msg

from goldenverba.components.document import Document, create_document
from goldenverba.components.interfaces import Reader
from goldenverba.server.types import FileConfig

# Optional imports with error handling
try:
    from pypdf import PdfReader
except ImportError:
    msg.warn("pypdf not installed, PDF functionality will be limited.")
    PdfReader = None

try:
    import spacy
except ImportError:
    msg.warn("spacy not installed, NLP functionality will be limited.")
    spacy = None

try:
    import docx
except ImportError:
    msg.warn("python-docx not installed, DOCX functionality will be limited.")
    docx = None


try:
    from openpyxl import load_workbook
except ImportError:
    msg.warn("openpyxl not installed, XLSX functionality will be limited.")
    load_workbook = None

class BasicReader(Reader):
    """
    The BasicReader reads text, code, PDF, and DOCX files.
    """

    def __init__(self):
        super().__init__()
        self.name = "Default"
        self.description = "Ingests text, code, PDF, and DOCX files"
        self.requires_library = ["pypdf", "docx", "spacy"]
        self.extension = [
            ".txt",
            ".py",
            ".js",
            ".html",
            ".css",
            ".md",
            ".mdx",
            ".json",
            ".pdf",
            ".docx",
            ".pptx",
            ".xlsx",
            ".csv",
            ".ts",
            ".tsx",
            ".vue",
            ".svelte",
            ".astro",
            ".php",
            ".rb",
            ".go",
            ".rs",
            ".swift",
            ".kt",
            ".java",
            ".c",
            ".cpp",
            ".h",
            ".hpp",
        ]  # Add supported text extensions

        # Initialize spaCy model if available
        self.nlp = spacy.blank("en") if spacy else None
        if self.nlp:
            self.nlp.add_pipe("sentencizer", config={"punct_chars": None})

    async def load(self, config: dict, fileConfig: FileConfig) -> list[Document]:
        """
        Load and process a file based on its extension.
        """
        msg.info(f"Loading {fileConfig.filename} ({fileConfig.extension.lower()})")

        if fileConfig.extension != "":
            decoded_bytes = base64.b64decode(fileConfig.content)

        try:
            if fileConfig.extension == "":
                file_content = fileConfig.content
            elif fileConfig.extension.lower() == "json":
                return await self.load_json_file(decoded_bytes, fileConfig)
            elif fileConfig.extension.lower() == "pdf":
                file_content = await self.load_pdf_file(decoded_bytes)
            elif fileConfig.extension.lower() == "docx":
                file_content = await self.load_docx_file(decoded_bytes)
            elif fileConfig.extension.lower() == "xlsx":
                file_content = await self.load_xlsx_file(decoded_bytes)
            elif fileConfig.extension.lower() in [
                ext.lstrip(".") for ext in self.extension
            ]:
                file_content = await self.load_text_file(decoded_bytes)
            else:
                try:
                    file_content = await self.load_text_file(decoded_bytes)
                except Exception as e:
                    raise ValueError(
                        f"Unsupported file extension: {fileConfig.extension}"
                    )

            return [create_document(file_content, fileConfig)]
        except Exception as e:
            msg.fail(f"Failed to load {fileConfig.filename}: {str(e)}")
            raise

    async def load_text_file(self, decoded_bytes: bytes) -> str:
        """Load and decode a text file."""
        try:
            return decoded_bytes.decode("utf-8")
        except UnicodeDecodeError:
            # Fallback to latin-1 if UTF-8 fails
            return decoded_bytes.decode("latin-1")

    async def load_json_file(
        self, decoded_bytes: bytes, fileConfig: FileConfig
    ) -> list[Document]:
        """Load and parse a JSON file."""
        try:
            json_obj = json.loads(decoded_bytes.decode("utf-8"))
            document = Document.from_json(json_obj, self.nlp)
            return (
                [document]
                if document
                else [create_document(json.dumps(json_obj, indent=2), fileConfig)]
            )
        except json.JSONDecodeError as e:
            raise ValueError(f"Invalid JSON in {fileConfig.filename}: {str(e)}")

    async def load_pdf_file(self, decoded_bytes: bytes) -> str:
        
        if not os.path.exists("img"):
            os.makedirs("img")
        pdf_bytes = io.BytesIO(decoded_bytes)
        doc = fitz.open(stream=pdf_bytes, filetype="pdf")
        full_content = []

        for page_num in range(len(doc)):
            page = doc.load_page(page_num)
           
            text_blocks = page.get_text("blocks")
            valid_text_blocks = []
            for block in text_blocks:
                if len(block) >= 5:  
                    x0, y0, x1, y1, text, *_ = block
                    if text.strip(): 
                        valid_text_blocks.append(block)

            image_list = page.get_images(full=True)
            images_with_coords = []
            for img_index, img in enumerate(image_list):
                xref = img[0]
                base_image = doc.extract_image(xref)
                image_bytes = base_image["image"]
                image_ext = base_image["ext"]

                try:
                    image_rect = page.get_image_bbox(img)
                    if image_rect:  
                        images_with_coords.append({
                            "index": img_index,
                            "rect": image_rect,
                            "bytes": image_bytes,
                            "ext": image_ext
                        })
                except Exception as e:
                    print(f"Error al obtener coordenadas de la imagen {img_index}: {e}")

           
            all_elements = []
            for block in valid_text_blocks:
                x0, y0, x1, y1, text, *_ = block
                all_elements.append(("text", (y0, x0, y1, x1, text)))  

            for img in images_with_coords:
                rect = img["rect"]
                all_elements.append(("image", (rect[1], rect[0], rect[3], rect[2], img)))  

            
            try:
                all_elements.sort(key=lambda elem: (elem[1][0], elem[1][1]))  # Ordenar por (y0, x0)
            except Exception as e:
                print(f"Error al ordenar elementos: {e}")
                all_elements = []  
            try:
                for elem_type, elem in all_elements:
                    if elem_type == "text":
                        _, _, _, _, text = elem
                        if text.strip():  
                            full_content.append(f" {text}")
                    elif elem_type == "image":
                        _, _, _, _, img = elem
                        img_index = img["index"]
                        image_bytes = img["bytes"]
                        image_ext = img["ext"]

                        image_pil = Image.open(io.BytesIO(image_bytes))

                        if image_pil.mode == "RGBA":
                            image_pil = image_pil.convert("RGB")

                        unique_id = uuid.uuid4().hex 
                        image_name = f"img/{unique_id}.{image_ext}" 

                        image_pil.save(image_name)

                        full_content.append(f"Imagen {page_num + 1}:\n{image_name}")
            except Exception as e:
                print(f"Error al procesar elementos: {e}")

        return "\n\n".join(full_content)

    async def load_docx_file(self, decoded_bytes: bytes) -> str:
        """Load and extract text from a DOCX file."""
        if not docx:
            raise ImportError(
                "python-docx is not installed. Cannot process DOCX files."
            )
        docx_bytes = io.BytesIO(decoded_bytes)
        reader = docx.Document(docx_bytes)
        return "\n".join(paragraph.text for paragraph in reader.paragraphs)

    async def load_xlsx_file(self, decoded_bytes: bytes) -> str:
        """Load and extract text from an XLSX file."""
        if not load_workbook:
            raise ImportError("openpyxl is not installed. Cannot process XLSX files.")
        xlsx_bytes = io.BytesIO(decoded_bytes)
        workbook = load_workbook(xlsx_bytes)
        text_content = []
        for sheet in workbook:
            for row in sheet.iter_rows(values_only=True):
                text_content.append("\t".join(str(cell) if cell is not None else "" for cell in row))
        return "\n".join(text_content)